
'''
Bot para telegram
'''
from fileinput import filename
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters)
import mysql.connector 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, Series, PieChart
from openpyxl.chart.series import DataPoint
from telegram.ext.callbackcontext import CallbackContext

db=mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    port=3306,
    database="bot",
)
def start(update, context: CallbackContext):

	print(update.message.from_user.id)

	context.bot.send_message(update.message.chat_id, "Hola, bienvenido, llevare un registro de tus ingresos y gastos, presiona el siguiente comando para comenzar /help.")

def help(update, context):
	
	context.bot.send_message(update.message.chat_id, "------------BIENVENIDOS----------- \n *************Comandos*************  \n ***************************************** \n /ingresar_salario Podras guardar en tu base de datos tus ingresos, ingresa la cantidad, descripcion, fecha y nombre del usuario en seguida del comando correspondiente de la siguiente manera: /ingresar_salario 10000000 comision 2022-04-02 greisy \n ***************************************** \n /ingresar_gasto Podras guardar en tu base de datos tus gastos, ingresa la cantidad, descripcion, fecha y nombre del usuario en seguida del comando correspondiente de la siguiente manera: /ingresar_gasto 1000 shampoo 2022-06-01 Oneyda \n ***************************************** \n /ingresos_personalizado Permitira saber todos los ingresos del mes actual. \n ***************************************** \n ***************************************** \n /gastos_personalizados Podras ver todos los gastos que has realizado. \n ***************************************** \n /excel_ingreso crea un archivo excel con los ingresos \n ***************************************** \n /excel_gasto crea un archivo excel con los gastos \n ***************************************** \n /grafica_barras_ingreso Mostrara un excel con un grafico de los ingresos que has realizado \n *****************************************\n /grafica_barras_gastos Mostrara un excel con un grafico de los gastos que has realizado. \n ***************************************** \n /grafica_pie_ingreso Mostrara un excel con un grafico de los ingresos que has realizado. \n ********************************** \n /grafica_pie_gastos Mostrara un excel con un grafico de los gastos que has realizado.")

def ingresar_salario(update,context):
  
	user = update.message.from_user.id
	cantidad = int(context.args[0])
	fecha = context.args[1]
	descripcion=context.args[2]
	nombre_usuario=context.args[3]
	cursor = db.cursor()
	cursor.execute("INSERT INTO ingresos (id_user,salario,fecha,descripcion,nombre_usuario)VALUES (%s,%s,%s,%s,%s)", (user,cantidad,descripcion,fecha,nombre_usuario))
	db.commit()

def ingresar_gasto(update,context):
	cantidad = int(context.args[0])
	fecha = context.args[1]
	descripcion=context.args[2]
	nombre_usuario=context.args[3]
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("INSERT INTO gastos (id_user,salario,fecha,descripcion,nombre_usuario)VALUES (%s,%s,%s,%s,%s)", (user,cantidad,descripcion,fecha,nombre_usuario))
	db.commit()

def ingresos_personalizado(update, context):

	fecha1 = context.args[0]
	fecha2 = context.args[1]
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("SELECT  salario,descripcion,fecha, nombre_usuario from ingresos WHERE fecha>=%s AND fecha<=%s AND id_user=%s",(fecha1,fecha2,user,))
	salario= cursor.fetchall()

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"
	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]
		sheet[f'D{i+2}'] = value[3]
	book.save('./static/ingresos/ingresos_mes.xlsx')		
	context.bot.send_document(update.message.chat_id,document=open("./static/ingresos/ingresos_mes.xlsx","rb"),filename="ingresos_mes.xlsx") 	

def gastos_personalizados(update,context):
	user = update.message.from_user.id
	cursor=db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha,nombre_usuario FROM gastos where id_user=%s",(user,))

	mes=cursor.fetchall()
	texto=""
	for database in mes:
    	
		texto = texto + '\n'+str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)
	context.bot.send_message(update.message.chat_id,"Estos son los gastos de este mes  "+texto)    		
    	
def excel_ingreso(update,context):
	user = update.message.from_user.id
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha,nombre_usuario FROM ingresos where id_user=%s",(user,))
	salario= cursor.fetchall()

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Falario"

	sheet['D1'] = "Nombre del Usuario"
	for i,value in enumerate(salario):

		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]
		sheet[f'D{i+2}'] = value[3]

	book.save('./static/ingresos/ingresos.xlsx')    	
	context.bot.send_document(update.message.chat_id,document=open("./static/ingresos/ingresos.xlsx","rb"),filename="ingresos.xlsx") 

def excel_gasto(update,context):
	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha, nombre_usuario FROM gastos where id_user=%s",(user,))
	salario= cursor.fetchall()
	texto=""
	for database in salario:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"
	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
		sheet[f'D{i+2}'] = value[3]
	
	book.save('./static/gastos/gastos.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("./static/gastos/gastos.xlsx","rb"),filename="gastos.xlsx")    	

#----------------------------------graficas de barra-------------------------------------

def grafica_barras_ingreso(update,context):

	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha, nombre_usuario FROM ingresos where id_user=%s",(user,))
	salario= cursor.fetchall()
	texto=""
	for database in salario:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"

	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
		sheet[f'D{i+2}'] = value[3]

	texto = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)

	grafica= BarChart()
	grafica.title = 'Gráfica de Ingresos'
	grafica.y_axis.title = 'eje Y'
	grafica.x_axis.title = 'eje X'
	grafica.add_data(texto)
	sheet.add_chart(grafica, "E15")
	
	book.save('grafica_barras_ingreso.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("grafica_barras_ingreso.xlsx","rb"),filename="grafica_barras_ingreso.xlsx")  	

def grafica_barras_gastos(update,context):
	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha, nombre_usuario FROM gastos where id_user=%s",(user,))
	salario= cursor.fetchall()
	texto=""
	for database in salario:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"

	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
		sheet[f'D{i+2}'] = value[3]

	texto = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)

	grafica= BarChart()
	grafica.title = 'Gráfica de Gastos'
	grafica.y_axis.title = 'eje Y'
	grafica.x_axis.title = 'eje X'
	grafica.add_data(texto)
	sheet.add_chart(grafica, "E15")
	
	book.save('grafica_barras_gastos.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("grafica_barras_gastos.xlsx","rb"),filename="grafica_barras_gastos.xlsx")

#---------------------------------grafica de pie------------------------------------------

def grafica_pie_ingreso(update,context):
	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha, nombre_usuario FROM ingresos where id_user=%s",(user,))
	salario= cursor.fetchall()
	texto=""
	for database in salario:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"

	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
		sheet[f'D{i+2}'] = value[3]

	texto = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)

	grafica= PieChart()
	grafica.title = 'Gráfica de Ingresos'
	grafica.add_data(texto)
	sheet.add_chart(grafica, "E15")
	
	book.save('grafica_pie_ingreso.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("grafica_pie_ingreso.xlsx","rb"),filename="grafica_pie_ingreso.xlsx")  	

def grafica_pie_gastos(update,context):
	user = update.message.from_user.id  	
	cursor = db.cursor()
	cursor.execute("SELECT salario,descripcion,fecha, nombre_usuario FROM gastos where id_user=%s",(user,))
	salario= cursor.fetchall()
	texto=""
	for database in salario:
        	
		texto =str(database[0])+" "+str(database[1])+" "+str(database[2])
		print(database)

	book = Workbook()
	sheet = book.active
	
	sheet['B1'] = 'Descripcion'

	sheet['C1'] = 'Fecha'

	sheet['A1'] = "Salario"

	sheet['D1'] = "Nombre del Usuario"

	for i,value in enumerate(salario):
		sheet[f'A{i+2}'] = value[0]
		sheet[f'B{i+2}'] = value[1]
		sheet[f'C{i+2}'] = value[2]	
		sheet[f'D{i+2}'] = value[3]

	texto = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)

	grafica= PieChart()
	grafica.title = 'Gráfica de Gastos'
	grafica.add_data(texto)
	sheet.add_chart(grafica, "E15")
	
	book.save('grafica_pie_gastos.xlsx')
	context.bot.send_document(update.message.chat_id,document=open("grafica_pie_gastos.xlsx","rb"),filename="grafica_pie_gastos.xlsx")


#-----------------------------------------------------------------------------------------
def main():
	TOKEN="5291069677:AAFB9wHXxjQDHIsa44TlosIv5gGCedptK_Y"
	updater=Updater(TOKEN, use_context=True)
	dp=updater.dispatcher

	#activarán el bot.
	dp.add_handler(CommandHandler('start',	start))
	dp.add_handler(CommandHandler('help',	help))
	dp.add_handler(CommandHandler('ingresar_salario',	ingresar_salario))
	dp.add_handler(CommandHandler('ingresar_gasto',	ingresar_gasto))
	dp.add_handler(CommandHandler('ingresos_personalizado',	ingresos_personalizado))
	dp.add_handler(CommandHandler('gastos_personalizados', gastos_personalizados))
	dp.add_handler(CommandHandler('excel_ingreso',	excel_ingreso))
	dp.add_handler(CommandHandler('excel_gasto',	excel_gasto))
	dp.add_handler(CommandHandler('grafica_barras_ingreso', grafica_barras_ingreso))
	dp.add_handler(CommandHandler('grafica_barras_gastos', grafica_barras_gastos))
	dp.add_handler(CommandHandler('grafica_pie_ingreso', grafica_pie_ingreso))
	dp.add_handler(CommandHandler('grafica_pie_gastos', grafica_pie_gastos))
	
	updater.start_polling()

	updater.idle()

if __name__ == '__main__':
	main()


